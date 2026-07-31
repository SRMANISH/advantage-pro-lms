"""Student-list import: parse a CSV/XLSX, validate every row, all-or-nothing.

If a single row is invalid the whole upload is rejected with a row-by-row error
report and nothing is written. A clean list imports inside one atomic transaction.
"""

from __future__ import annotations

import csv
import io
import re
import zipfile

from django.conf import settings
from django.core.exceptions import ValidationError as DjangoValidationError
from django.core.validators import validate_email
from django.db import transaction

from accounts.models import User, UserStatus
from batches.models import Batch
from core.roles import Role

from .models import Enrollment

# Faculty assignment is done in the batch UI (conflict-checked), not per-student at import.
REQUIRED = ["registration_number", "name", "email", "phone", "batch", "course"]
OPTIONAL = ["address", "guardian", "employment_company"]

_PHONE_RE = re.compile(r"^\+?\d[\d\s-]{6,18}$")

# "Registration ID" is the student recognition term. Accept friendly spellings for that
# column (and tolerate case/spacing on every column) so admins can use "Registration ID"
# in the sheet while we still store it as the canonical ``registration_number``.
_HEADER_ALIASES = {
    "registration_id": "registration_number",
    "registration_no": "registration_number",
    "registration_number": "registration_number",
    "reg_id": "registration_number",
    "reg_no": "registration_number",
}


def _canonical_header(key: str) -> str:
    norm = re.sub(r"\s+", "_", (key or "").strip().lower())
    return _HEADER_ALIASES.get(norm, norm)


def parse_rows(uploaded) -> list[dict]:
    """Return a list of dict rows from a CSV or XLSX upload. Raises ValueError on bad files.

    The size checks happen *before* parsing. MAX_IMPORT_ROWS is enforced by the view on the
    parsed result, which is too late to help: by then the file has already been decompressed
    and materialised in memory, which is exactly what a hostile upload is trying to make us
    do. This is the largest untrusted-input surface in the application — an MIS user uploads
    a spreadsheet they were sent.
    """
    max_mb = int(getattr(settings, "MAX_IMPORT_UPLOAD_MB", 10))
    size = getattr(uploaded, "size", 0) or 0
    if size > max_mb * 1024 * 1024:
        raise ValueError(f"File is too large (maximum {max_mb} MB).")

    name = (uploaded.name or "").lower()
    raw = uploaded.read()
    # `size` can be absent or wrong on a file-like that is not an UploadedFile; the bytes we
    # actually hold are the authority.
    if len(raw) > max_mb * 1024 * 1024:
        raise ValueError(f"File is too large (maximum {max_mb} MB).")

    if name.endswith(".xlsx"):
        return _parse_xlsx(raw)
    if name.endswith(".csv") or not name:
        return _parse_csv(raw)
    raise ValueError("Unsupported file type. Upload a .csv or .xlsx file.")


def _reject_hostile_workbook(raw: bytes) -> None:
    """Inspect the XLSX container before handing it to openpyxl.

    An .xlsx is a ZIP. ``read_only=True`` streams rows, but it still decompresses
    sharedStrings.xml in full, so a small file can expand to gigabytes — the classic zip bomb.
    The ZIP central directory records each entry's decompressed size without decompressing
    anything, so the whole check costs one directory read.

    Two limits, because either alone is evadable: a total decompressed cap catches the big
    payload, and a per-entry compression ratio catches a file that stays under the cap while
    still being wildly disproportionate.
    """
    max_mb = int(getattr(settings, "MAX_IMPORT_DECOMPRESSED_MB", 200))
    max_ratio = int(getattr(settings, "MAX_IMPORT_COMPRESSION_RATIO", 200))
    max_sheets = int(getattr(settings, "MAX_IMPORT_SHEETS", 25))

    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as zf:
            infos = zf.infolist()
            total = sum(i.file_size for i in infos)
            if total > max_mb * 1024 * 1024:
                raise ValueError(
                    "This workbook expands to far more data than an enrolment list should "
                    "— it was not opened. Export a plain .csv and upload that instead."
                )
            for info in infos:
                if info.compress_size > 0 and info.file_size / info.compress_size > max_ratio:
                    raise ValueError(
                        "This workbook is compressed far beyond what a spreadsheet needs "
                        "— it was not opened. Export a plain .csv and upload that instead."
                    )
            sheets = sum(1 for i in infos if i.filename.startswith("xl/worksheets/"))
            if sheets > max_sheets:
                raise ValueError(
                    f"This workbook has {sheets} sheets. Upload one with the enrolment list "
                    "on a single sheet."
                )
    except zipfile.BadZipFile as exc:
        raise ValueError("That .xlsx file is not a readable workbook.") from exc


def _parse_csv(raw: bytes) -> list[dict]:
    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    if reader.fieldnames is None:
        raise ValueError("The file is empty.")
    return [{_canonical_header(k): (v or "").strip() for k, v in row.items()} for row in reader]


def _parse_xlsx(raw: bytes) -> list[dict]:
    from openpyxl import load_workbook

    _reject_hostile_workbook(raw)
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = [_canonical_header(str(h)) if h is not None else "" for h in next(rows)]
    except StopIteration as exc:
        raise ValueError("The file is empty.") from exc
    # Bail out during iteration rather than after: the view's MAX_IMPORT_ROWS check runs on a
    # fully-built list, so a sheet claiming a million rows would be materialised before being
    # rejected. A little headroom over the limit so the view still produces the friendly
    # "too many rows" message for merely-oversized (not hostile) files.
    hard_cap = int(getattr(settings, "MAX_IMPORT_ROWS", 5000)) * 2
    out: list[dict] = []
    for values in rows:
        if values is None or all(v is None for v in values):
            continue
        if len(out) > hard_cap:
            raise ValueError(
                "This sheet has far more rows than an enrolment list should — it was not "
                f"read past {hard_cap} rows."
            )
        cells = ["" if v is None else str(v).strip() for v in values]
        out.append(dict(zip(header, cells, strict=False)))
    return out


def validate_and_build(rows: list[dict]) -> tuple[list[dict], list[dict]]:
    """Return (errors, prepared). Errors is a row-by-row list; prepared is build specs."""
    errors: list[dict] = []
    if not rows:
        return [{"row": 0, "field": "file", "message": "No data rows found."}], []

    missing_cols = [c for c in REQUIRED if c not in rows[0]]
    if missing_cols:
        return [
            {"row": 1, "field": ", ".join(missing_cols), "message": "Missing required column(s)."}
        ], []

    existing_regs = set(
        User.objects.filter(
            username__in=[r.get("registration_number", "") for r in rows]
        ).values_list("username", flat=True)
    )
    batches = {b.code: b for b in Batch.objects.select_related("course").all()}

    seen_regs: set[str] = set()
    prepared: list[dict] = []

    for index, row in enumerate(rows, start=2):  # row 1 is the header

        def err(field, message, _row=index):
            errors.append({"row": _row, "field": field, "message": message})

        reg = row.get("registration_number", "").strip()
        name = row.get("name", "").strip()
        email = row.get("email", "").strip()
        phone = row.get("phone", "").strip()
        batch_code = row.get("batch", "").strip()
        course_code = row.get("course", "").strip()

        for field in REQUIRED:
            if not row.get(field, "").strip():
                err(field, "This field is required.")

        if reg:
            if reg in existing_regs:
                err("registration_number", "A student with this Registration ID already exists.")
            if reg in seen_regs:
                err("registration_number", "Duplicate Registration ID within the file.")
            seen_regs.add(reg)

        if email:
            try:
                validate_email(email)
            except DjangoValidationError:
                err("email", "Invalid email address.")

        if phone and not _PHONE_RE.match(phone):
            err("phone", "Invalid phone number.")

        batch = batches.get(batch_code)
        if batch_code and batch is None:
            err("batch", f"Unknown batch '{batch_code}'.")
        if batch and course_code and batch.course.code != course_code:
            err("course", f"Course '{course_code}' does not match batch's course.")

        prepared.append(
            {
                "registration_number": reg,
                "name": name,
                "email": email,
                "phone": phone,
                "batch": batch,
                "address": row.get("address", "").strip(),
                "guardian": row.get("guardian", "").strip(),
                "employment_company": row.get("employment_company", "").strip(),
            }
        )

    if errors:
        return errors, []
    return [], prepared


@transaction.atomic
def do_import(prepared: list[dict]) -> list[User]:
    """Create student accounts + enrolments atomically. Students start as 'pending'."""
    created: list[User] = []
    for spec in prepared:
        student = User.objects.create_user(
            username=spec["registration_number"],
            password=None,  # set during two-step account setup
            role=Role.STUDENT,
            status=UserStatus.PENDING,
            full_name=spec["name"],
            email=spec["email"],
            phone=spec["phone"],
        )
        Enrollment.objects.create(
            student=student,
            batch=spec["batch"],
            registration_number=spec["registration_number"],
            address=spec["address"],
            guardian=spec["guardian"],
            employment_company=spec["employment_company"],
        )
        created.append(student)
    return created
